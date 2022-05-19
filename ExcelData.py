from openpyxl import Workbook, load_workbook
import common
import const
from Model import Variable, VariableValue, TaskType, Technique, Relationship, Group, Student, Level, LocalInsInfo

class ExcelData :

    def __init__(self, url=""):
        self.url = url

    # function is check column A is correct
    def checkColumnAIsCorrect(self, sheetname):
        try:
            workBook = load_workbook(self.url)
            workSheet = workBook[sheetname]
            row = 0
            flg = True
            lstRow = []
            # Check column A is not empty cell in between 2 cell
            for cell in workSheet['A']:
                row = row + 1
                if (cell.value is None) :
                    if (row + 1 < (len(workSheet['A']) - 1)):
                        if (workSheet['A'][row + 1] is not None and workSheet['A'][row + 1] != "") :
                            lstRow.append(str(row))
                            flg = False
                    else:
                        lstRow.append(str(row))
                        flg = False
            workBook.close()
            if not flg :
                common.writeNotification(const.MSG_001 , [sheetname, ",".join(lstRow)])
            return flg
        except Exception as e:
            common.writeLog("Validate > checkColumnANotEmpty() " + (e))

    # function is check reference of sheet Task Type
    def checkReferenceTaskType(self):
        result = False
        try :
            wb = load_workbook(self.url)
            cellFirstSubTask = ""
            cellLastSubTask =""

            # Check number of column variabel's tasktype match with code of variable
            cntCol = 0
            checkIsCellSubTask = False
            ws = wb[const.SHEET_NAME_VARIABLE]
            cntRow = self.countRowColumn(ws)

            ws = wb[const.SHEET_NAME_TASK_TYPE]
            for col in ws[1]:
                if (col.value == const.SUB_TASK) :
                    cellFirstSubTask = col.coordinate
                    checkIsCellSubTask = True
                if (col.value is None and not checkIsCellSubTask) :
                    cntCol = cntCol + 1
                cellLastSubTask = col.coordinate

            if cntCol == cntRow - 1 :
                result = True
            else:
                common.writeNotification(const.MSG_002)

            # check code sub task type is exits in task type
            lstAllTaskType = self.getValuesColumn(wb[const.SHEET_NAME_TASK_TYPE], 'A')
            lstAllTaskType = lstAllTaskType if lstAllTaskType != False else []
            cellFirstSubTask = cellFirstSubTask.replace("1","")
            cellLastSubTask = cellLastSubTask.replace("1","")
            lstSubTask = []

            for row in (ws[cellFirstSubTask + ":" + cellLastSubTask]):
                for cell in row :
                    if not None and cell.value not in lstSubTask:
                        lstSubTask.append(cell.value)
            lstDiff = list(set(lstSubTask) - set(lstAllTaskType))
            lstDiff.remove(const.SUB_TASK)
            lstDiff.remove(None)

            if (len(lstDiff) > 0):
                result = False
                common.writeNotification(const.MSG_003, [",".join(lstDiff)])
            wb.close()
        except Exception as e:
            common.writeLog("Validate > checkReferenceVariableValue() " + (e))

        return result

    # function is check reference of sheet Variable Value
    def checkReferenceVariableValue(self):
        result = False
        try :
            wb = load_workbook(self.url)

            #check code variable of variable value isn't match with code variable of sheet variable
            lstVariablesOfVariableSheet = self.getValuesColumn(wb[const.SHEET_NAME_VARIABLE], 'A')
            lstVariableOfVariableValueSheet = self.getValuesColumn(wb[const.SHEET_NAME_VARIABLE_VALUE], 'A')
            listDiff = list(set(lstVariablesOfVariableSheet) - set(lstVariableOfVariableValueSheet))
            listDiff = [i for i in listDiff if i]

            if (len(listDiff) == 0) :
                result = True
            else:
                common.writeNotification(const.MSG_004, [",".join(listDiff)])

            wb.close()
        except Exception as e:
            common.writeLog("Validate > checkReferenceVariableValue() " + str(e))

        return result

    # function support convert a value cell of sheet Technique to list (ex : cell A1 has value "TE001,TE003,TE004")
    def convertCellTechniqueSheet(self, lst = []):
        result = []
        if len(lst) > 0 :
            for cell in lst :
                if len(str(cell).split(",")) > 1 :
                    for item in str(cell).split(",") :
                        if item.strip() != "" : result.append(item.strip())
                else :
                    result.append(str(cell).strip())

        return result

    # function is check reference of sheet technique
    def checkReferenceTechnique(self):
        result = False
        try :
            wb = load_workbook(self.url)

            # check code task type of sheet Task Type is exist in code of sheet Task Type
            lstAllTTofSheetTaskType = self.getValuesColumn(wb[const.SHEET_NAME_TASK_TYPE], 'A')
            lstAllPPragOfSheetTechnique = self.getValuesColumn(wb[const.SHEET_NAME_TECHNIQUE], 'C')
            lstAllPPragOfSheetTechnique = [i for i in lstAllPPragOfSheetTechnique if i]
            lstAllTTofSheetTaskType = [i for i in lstAllTTofSheetTaskType if i]

            lstDiff = list(set(lstAllTTofSheetTaskType if not lstAllTTofSheetTaskType else []) - set(lstAllPPragOfSheetTechnique if not lstAllPPragOfSheetTechnique else []))
            if len(lstDiff) == 0 :
                result = True
            else :
                result = False
                common.writeNotification(const.MSG_005, {",".join(lstDiff)})

            # check code technique of sub technique in sheet Technique
            lstAllCodeTechnique = self.getValuesColumn(wb[const.SHEET_NAME_TECHNIQUE], 'A')
            lstAllCodeTechnique = [i for i in lstAllCodeTechnique if i] if lstAllCodeTechnique != False else []

            # check for concurency global
            lstConcuGl = self.getValuesColumn(wb[const.SHEET_NAME_TECHNIQUE], 'D')
            lstConcuGl = [i for i in lstConcuGl if i] if lstConcuGl != False else []
            lstConcuGl = self.convertCellTechniqueSheet(lstConcuGl)
            lstDiffConCuGl = list(set(lstConcuGl) - set(lstAllCodeTechnique))
            lstDiffConCuGl = list(filter('None', lstDiffConCuGl))
            if len(lstDiffConCuGl) == 0 :
                result = True
            else:
                result = False
                common.writeNotification(const.MSG_006, [",".join(lstDiffConCuGl)])

            # check for concurency inc
            lstConcuInc = self.getValuesColumn(wb[const.SHEET_NAME_TECHNIQUE], 'E')
            lstConcuInc = [i for i in lstConcuInc if i] if lstConcuInc != False else []
            lstConcuInc = self.convertCellTechniqueSheet(lstConcuInc)
            lstDiffConCuInc = list(set(lstConcuInc) - set(lstAllCodeTechnique))
            lstDiffConCuInc = list(filter('None', lstDiffConCuInc))
            if len(lstDiffConCuInc) == 0 :
                result = True
            else:
                result = False
                common.writeNotification(const.MSG_007, [",".join(lstDiffConCuInc)])

            # check for concurency ext
            lstConcuExt = self.getValuesColumn(wb[const.SHEET_NAME_TECHNIQUE], 'F')
            lstConcuExt = [i for i in lstConcuExt if i] if lstConcuExt != False else []
            lstConcuExt = self.convertCellTechniqueSheet(lstConcuExt)
            lstDiffConCuExt = list(set(lstConcuExt) - set(lstAllCodeTechnique))
            lstDiffConCuExt = list(filter('None', lstDiffConCuExt))
            if len(lstDiffConCuExt) == 0 :
                result = True
            else:
                result = False
                common.writeNotification(const.MSG_008, [",".join(lstDiffConCuExt)])

            wb.close()
        except Exception as e:
            common.writeLog("Validate > checkReferenceTechnique() " + str(e))

        return result

    #  function is check reference of sheet student
    def checkReferenceStudent(self):
        result = True
        try :
            wb = load_workbook(self.url)
            tmp = []
            # check code of group in sheet student has exist code in sheet Group
            lstCdGroupSheetStudent = self.getValuesColumn(wb[const.SHEET_NAME_STUDENT], 'C')
            lstCdGroupSheetGroup = self.getValuesColumn(wb[const.SHEET_NAME_GROUP], 'A')

            for i in lstCdGroupSheetStudent:
                if i not in lstCdGroupSheetGroup :
                    tmp.append(i)
                    result = False
            wb.close()
            if len(tmp) > 0 :
                common.writeNotification(const.MSG_009, {",".join(tmp)})
            return result
        except Exception as e:
            common.writeLog("Validate > checkReferenceStudent() " + str(e))

    # function is check reference of sheet LocalInsInfo
    def checkReferenceLocalInsInfo(self):
        result = True
        try:
            wb = load_workbook(self.url)
            lstCdStudent = []
            # check has code student is exist in sheet Student
            lstCdStudentSheetStudent = self.getValuesColumn(wb[const.SHEET_NAME_STUDENT], 'A')
            lstCdStudentSheetLocalInsInfo = self.getValuesColumn(wb[const.SHEET_NAME_LOCALINSINFO], 'B')
            i = 2
            for cdStudent in lstCdStudentSheetLocalInsInfo :
                if cdStudent not in lstCdStudentSheetStudent :
                    result = False
                    lstCdStudent.append("B" + str(i))
                i = i + 1
            if len(lstCdStudent) > 0 :
                common.writeNotification(const.MSG_010, {",".join(lstCdStudent)})

            # check has code level is exist in sheet Level
            lstLevelSheetLevel = self.getValuesColumn(wb[const.SHEET_NAME_LEVEL], 'B')
            lstLevelSheetLocalInsInfo = self.getValuesColumn(wb[const.SHEET_NAME_LOCALINSINFO], 'D')
            lstLevel = []
            i = 2
            for level in lstLevelSheetLocalInsInfo :
                if level.lower() not in [x.lower() for x in lstLevelSheetLevel]:
                    result = False
                    lstLevel.append("D" + str(i))
                i = i + 1
            if len(lstLevel) > 0 :
                common.writeNotification(const.MSG_011, {",".join(lstLevel)})

            # check has code technique is exist in sheet technique
            lstCdTechiqueSheetTechnique = self.getValuesColumn(wb[const.SHEET_NAME_TECHNIQUE], 'A')
            lstCdTechniqeSheetLocalInsInfo = self.getValuesColumn(wb[const.SHEET_NAME_LOCALINSINFO], 'C')
            i = 2
            lstTech = []
            for techs in lstCdTechniqeSheetLocalInsInfo:
                if len(techs.split(',')) > 1 :
                    lstCdTech = techs.replace(" ","").split(',')
                    for tech in lstCdTech :
                        if tech not in lstCdTechiqueSheetTechnique :
                            result = False
                            lstTech.append("C" + str(i))
                else :
                    if techs not in lstCdTechiqueSheetTechnique:
                        result = False
                        lstTech.append("C" + str(i))

                i = i + 1
            if len(lstTech) > 0 :
                common.writeNotification(const.MSG_012, {",".join(lstTech)})

            wb.close()
            return result
        except Exception as e:
            common.writeLog("Validate > checkReferenceLocalInsInfo() " + str(e))

    # function count row of column
    def countRowColumn(self, workSheet):
        cnt = 0;
        for row in workSheet:
            cnt = cnt + 1
        return cnt

    # function is get all values of a column
    def getValuesColumn(self, ws, columnName):
        if columnName.strip() != "" :
            lst =[]
            for col in ws[columnName]:
                lst.append(str(col.value).strip())
            del lst[0]
            return lst
        else :
            return False

    # function validate excel is correct template
    # return : true if validate or false if wrong
    def validateFileExcel(self):
        wb = load_workbook(self.url, read_only=True)

        # case check file excel is not exist 8 sheet Variable Value, Variable, Technique, Task Type, Level, Group, Student, LocalInsInfo
        if (const.SHEET_NAME_VARIABLE not in wb.sheetnames
                or const.SHEET_NAME_VARIABLE not in wb.sheetnames
                or const.SHEET_NAME_TASK_TYPE not in wb
                or const.SHEET_NAME_TECHNIQUE not in wb.sheetnames
                or const.SHEET_NAME_LEVEL not in wb.sheetnames
                or const.SHEET_NAME_GROUP not in wb.sheetnames
                or const.SHEET_NAME_STUDENT not in wb.sheetnames
                or const.SHEET_NAME_LOCALINSINFO not in wb.sheetnames):
            return False
        else:
            # case check in 8 sheet is exist column ID (column A)
            resultTechnique = self.checkColumnAIsCorrect(const.SHEET_NAME_TECHNIQUE)
            resultVariableValue = self.checkColumnAIsCorrect(const.SHEET_NAME_VARIABLE_VALUE)
            resultVariable = self.checkColumnAIsCorrect(const.SHEET_NAME_VARIABLE)
            resultTaskType = self.checkColumnAIsCorrect(const.SHEET_NAME_TASK_TYPE)
            resultLevel = self.checkColumnAIsCorrect(const.SHEET_NAME_LEVEL)
            resultGroup = self.checkColumnAIsCorrect(const.SHEET_NAME_GROUP)
            resultLocalInsInfo = self.checkColumnAIsCorrect(const.SHEET_NAME_LOCALINSINFO)
            resultStudent = self.checkColumnAIsCorrect(const.SHEET_NAME_STUDENT)
            if (resultTechnique == True and resultTaskType  == True and resultVariable == True and resultVariableValue == True
            and resultStudent == True and resultGroup == True and resultLocalInsInfo == True):
                referenceTaskType = self.checkReferenceTaskType()
                referenceVariableValue = self.checkReferenceVariableValue()
                referenceTechnique = self.checkReferenceTechnique()
                referenceStudent = self.checkReferenceStudent()
                referenceLocalInsInfo = self.checkReferenceLocalInsInfo()
                if (referenceTechnique == True and referenceTaskType == True and referenceVariableValue == True
                and referenceStudent == True and referenceLocalInsInfo == True) :
                    return True
                else :
                    return False
            else:
                return False

    # function get data node of sheet Variable
    def getDataSheetVariable(self):
        try :
            lstVariable = []
            wb = load_workbook(self.url)
            lstCdVariable = self.getValuesColumn(wb[const.SHEET_NAME_VARIABLE], 'A')
            lstCdVariable = lstCdVariable if lstCdVariable != False else []
            lstValueVariable = self.getValuesColumn(wb[const.SHEET_NAME_VARIABLE], 'B')
            lstValueVariable = lstValueVariable if lstValueVariable != False else []

            for i in range(len(lstCdVariable)):
                variable = Variable(lstCdVariable[i], lstValueVariable[i], const.SHEET_NAME_VARIABLE)
                lstVariable.append(variable)
            wb.close()
            return lstVariable
        except Exception as e :
            common.writeLog("ExcelData > getDataSheetVariable() " + str(e))

    # function get data node of sheet Variable Value
    def getDataSheetVariableValue(self):
        try:
            lstVariableValue = []
            wb = load_workbook(self.url)

            lstCdVariableValue = self.getValuesColumn(wb[const.SHEET_NAME_VARIABLE_VALUE], 'B')
            lstCdVariableValue = lstCdVariableValue if lstCdVariableValue != False else []

            lstTypeVariableValue = self.getValuesColumn(wb[const.SHEET_NAME_VARIABLE_VALUE], 'C')
            lstTypeVariableValue = lstTypeVariableValue if lstTypeVariableValue != False else []

            lstIsValueUsedVariableValue = self.getValuesColumn(wb[const.SHEET_NAME_VARIABLE_VALUE], 'D')
            lstIsValueUsedVariableValue = lstIsValueUsedVariableValue if lstIsValueUsedVariableValue != False else []

            lstValueVariableValue = self.getValuesColumn(wb[const.SHEET_NAME_VARIABLE_VALUE], 'E')
            lstValueVariableValue = lstValueVariableValue if lstValueVariableValue != False else []

            for i in range(len(lstCdVariableValue)):
                variableValue = VariableValue(lstCdVariableValue[i],lstValueVariableValue[i], lstTypeVariableValue[i], lstIsValueUsedVariableValue[i])
                lstVariableValue.append(variableValue)

            wb.close()
            return lstVariableValue
        except Exception as e:
            common.writeLog("ExcelData > getDataSheetVariableValue() " + str(e))

    # function get data node of sheet level
    def getDataSheetLevel(self):
        try:
            lstLevel = []
            wb = load_workbook(self.url)

            lstCdLevel = self.getValuesColumn(wb[const.SHEET_NAME_LEVEL], 'A')
            lstCdLevel = lstCdLevel if lstCdLevel != False else []

            lstValueLevel = self.getValuesColumn(wb[const.SHEET_NAME_LEVEL], 'B')
            lstValueLevel = lstValueLevel if lstValueLevel != False else []

            for i in range(len(lstCdLevel)):
                level = Level(lstCdLevel[i],lstValueLevel[i])
                lstLevel.append(level)

            wb.close()
            return lstLevel
        except Exception as e:
            common.writeLog("ExcelData > getDataSheetVariableValue() " + str(e))

    # function get data node of sheet Task Type
    def getDataSheetTaskType(self):
        try:
            lstTaskType = []
            wb = load_workbook(self.url)

            lstCdTaskType = self.getValuesColumn(wb[const.SHEET_NAME_TASK_TYPE], 'A')
            lstCdTaskType = lstCdTaskType if lstCdTaskType != False else []

            lstNameTaskType = self.getValuesColumn(wb[const.SHEET_NAME_TASK_TYPE], 'B')
            lstNameTaskType = lstNameTaskType if lstNameTaskType != False else []

            for i in range(len(lstCdTaskType)):
                taskType = TaskType(lstCdTaskType[i], lstNameTaskType[i])
                lstTaskType.append(taskType)

            wb.close()
            return lstTaskType
        except Exception as e:
            common.writeLog("ExcelData > getDataSheetTaskType() " + str(e))

    # function get data node of sheet Technique
    def getDataSheetTechnique(self):
        try:
            lstTechnique = []
            wb = load_workbook(self.url)

            lstCdTechnique = self.getValuesColumn(wb[const.SHEET_NAME_TECHNIQUE], 'A')
            lstCdTechnique = lstCdTechnique if lstCdTechnique != False else []

            lstValueTechnique = self.getValuesColumn(wb[const.SHEET_NAME_TECHNIQUE], 'B')
            lstValueTechnique = lstValueTechnique if lstValueTechnique != False else []

            for i in range(len(lstCdTechnique)):
                technique = Technique(lstCdTechnique[i], lstValueTechnique[i])
                lstTechnique.append(technique)

            wb.close()
            return lstTechnique
        except Exception as e:
            common.writeLog("ExcelData > getDataSheetTechnique() " + str(e))

    # function get data node of sheet Level
    def getDataSheetLevel(self):
        try:
            lstLevel = []
            wb = load_workbook(self.url)

            lstCdLevel = self.getValuesColumn(wb[const.SHEET_NAME_LEVEL], 'A')
            lstCdLevel = lstCdLevel if lstCdLevel != False else []
            lstValueLevel = self.getValuesColumn(wb[const.SHEET_NAME_LEVEL], 'B')
            lstValueLevel = lstValueLevel if lstValueLevel != False else []

            for i in range(len(lstCdLevel)):
                level = Level(lstCdLevel[i], lstValueLevel[i])
                lstLevel.append(level)

            wb.close()
            return lstLevel
        except Exception as e:
            common.writeLog("ExcelData > getDataSheetLevel() " + str(e))

    # function get data node of sheet Student
    def getDataSheetStudent(self):
        try:
            lstStudents = []
            wb = load_workbook(self.url)

            lstCdStudent = self.getValuesColumn(wb[const.SHEET_NAME_STUDENT], 'A')
            lstCdStudent = lstCdStudent if lstCdStudent != False else []

            lstValueStudent = self.getValuesColumn(wb[const.SHEET_NAME_STUDENT], 'B')
            lstValueStudent = lstValueStudent if lstValueStudent != False else []

            for i in range(len(lstCdStudent)):
                student = Student(lstCdStudent[i], lstValueStudent[i])
                lstStudents.append(student)

            wb.close()

            return lstStudents
        except Exception as e:
            common.writeLog("ExcelData > getDataSheetLevel() " + str(e))

    # function get data node of sheet Group
    def getDataSheetGroup(self):
        try:
            lstGroup= []
            wb = load_workbook(self.url)

            lstCdGroup = self.getValuesColumn(wb[const.SHEET_NAME_GROUP], 'A')
            lstCdGroup = lstCdGroup if lstCdGroup != False else []

            lstValueGroup = self.getValuesColumn(wb[const.SHEET_NAME_GROUP], 'B')
            lstValueGroup = lstValueGroup if lstValueGroup != False else []

            for i in range(len(lstCdGroup)):
                group = Group(lstCdGroup[i], lstValueGroup[i])
                lstGroup.append(group)

            wb.close()

            return lstGroup
        except Exception as e:
            common.writeLog("ExcelData > getDataSheetGroup() " + str(e))

    # function get data node of sheet LocalInsInfo
    def getDataSheetLocalInsInfo(self):
        try:
            lstLocalInsInfo= []
            wb = load_workbook(self.url)

            lstCdLocalInsInfo = self.getValuesColumn(wb[const.SHEET_NAME_LOCALINSINFO], 'A')
            lstCdLocalInsInfo = lstCdLocalInsInfo if lstCdLocalInsInfo != False else []

            for i in range(len(lstCdLocalInsInfo)):
                localInsInfo = LocalInsInfo(lstCdLocalInsInfo[i])
                lstLocalInsInfo.append(localInsInfo)

            wb.close()

            return lstLocalInsInfo
        except Exception as e:
            common.writeLog("ExcelData > getDataSheetLocalInsInfo() " + str(e))

    #function  support getRelationshipSheetTaskType() for task find code by value variable value
    def getCodeByValueVariableValue(self,lst , value):
        result = ""
        for item in lst :
            if item.value.strip() == value.strip() :
                result = item.code
        return result

    # function get data relationship of sheet Task Type
    def getRelationshipSheetTaskType(self):
        try :
            lstRelationship = []
            lstCol = []
            isSubTak = False
            dataSheetVariableValue = self.getDataSheetVariableValue()

            wb = load_workbook(self.url)
            lstAllCdTaskType = self.getValuesColumn(wb[const.SHEET_NAME_TASK_TYPE], 'A')

            # create list col name without code task type and name task type
            for col in wb[const.SHEET_NAME_TASK_TYPE][1] :
                if (col.value == const.SUB_TASK) :
                    isSubTak = True
                if isSubTak :
                    lstCol.append((col.coordinate + "s").replace("1", ""))
                else :
                    lstCol.append((col.coordinate).replace("1", ""))
            del lstCol[0:2]

            for col in lstCol :
                if "s" in col :
                    # create relationship reference task type with sub task type
                    lstSubTaskType = self.getValuesColumn(wb[const.SHEET_NAME_TASK_TYPE], col.replace("s",""))
                    for i in range(len(lstAllCdTaskType)) :
                        if lstSubTaskType[i] is not None :
                            relationship = Relationship(const.NAME_TASK_TYPE,lstAllCdTaskType[i], lstSubTaskType[i],const.NAME_TASK_TYPE, const.RELATIONSHIP_TASKTYPE_TASKTYPE)
                            lstRelationship.append(relationship)
                else :
                    # create relationship reference task type with variable value
                    lstVariableValue = self.getValuesColumn(wb[const.SHEET_NAME_TASK_TYPE], col)
                    for i in range(len(lstAllCdTaskType)) :
                        if lstVariableValue[i] is not  None :
                            codeVariableValue = self.getCodeByValueVariableValue(dataSheetVariableValue, lstVariableValue[i])
                            if codeVariableValue != "" :
                                relationship = Relationship(const.NAME_TASK_TYPE, lstAllCdTaskType[i], codeVariableValue,const.NAME_VARIABLE_VALUE, const.RELATIONSHIP_TASKTYPE_VARIABLEVALUE)
                                lstRelationship.append(relationship)
            wb.close()
            return lstRelationship
        except Exception as e :
            common.writeLog("Validate > getRelationshipSheetTaskType() " + str(e))

    # function get data relationship of sheet technique
    def getRelationshipSheetTechnique(self):
        try :
            lstRelationship = []
            wb = load_workbook(self.url)
            lstCdTechnique = self.getValuesColumn(wb[const.SHEET_NAME_TECHNIQUE], 'A')

            #create relationship technique with task type
            lstCdTaskType = self.getValuesColumn(wb[const.SHEET_NAME_TECHNIQUE], 'C')
            for i in range(len(lstCdTechnique)) :
                relationship = Relationship(const.NAME_TECHNIQUE, str(lstCdTechnique[i]).strip(),str(lstCdTaskType[i]).strip(),const.NAME_TASK_TYPE , const.RELATIONSHIP_TASKTYPE_TECHNIQUE)
                lstRelationship.append(relationship)

            # create relationship technique with global concurency
            lstTechniqueConcurencyGlobal = self.getValuesColumn(wb[const.SHEET_NAME_TECHNIQUE], 'D')
            for i in range(len(lstCdTechnique)) :
                if len(str(lstTechniqueConcurencyGlobal[i]).split(",")) > 1 :
                    for item in lstTechniqueConcurencyGlobal[i].split(",") :
                        if item != "" :
                            relationship = Relationship(const.NAME_TECHNIQUE, str(lstCdTechnique[i]).strip(), item.strip(),const.NAME_TECHNIQUE ,const.RELATIONSHIP_TT_GLOBAL)
                            lstRelationship.append(relationship)

                else :
                    if lstTechniqueConcurencyGlobal[i] is not None :
                        relationship = Relationship(const.NAME_TECHNIQUE, str(lstCdTechnique[i]).strip(), str(lstTechniqueConcurencyGlobal[i]).strip(),const.NAME_TECHNIQUE ,const.RELATIONSHIP_TT_GLOBAL)
                        lstRelationship.append(relationship)

            # create relationship technique with inc concurency
            lstTechniqueConcurencyInc = self.getValuesColumn(wb[const.SHEET_NAME_TECHNIQUE], 'E')
            for i in range(len(lstCdTechnique)):
                if len(str(lstTechniqueConcurencyInc[i]).split(",")) > 1:
                    for item in lstTechniqueConcurencyInc[i].split(","):
                        if item != "":
                            left = str(lstCdTechnique[i]).replace(" ","")
                            right = str(item).replace(" ","")
                            relationship = Relationship(const.NAME_TECHNIQUE, left, right ,const.NAME_TECHNIQUE , const.RELATIONSHIP_TT_INC)
                            lstRelationship.append(relationship)
                else:
                    if lstTechniqueConcurencyInc[i] is not None:
                        relationship = Relationship(const.NAME_TECHNIQUE,str(lstCdTechnique[i]).strip(), str(lstTechniqueConcurencyInc[i]).strip(),const.NAME_TECHNIQUE, const.RELATIONSHIP_TT_INC)
                        lstRelationship.append(relationship)

            # create relationship technique with ext concurency
            lstTechniqueConcurencyExt = self.getValuesColumn(wb[const.SHEET_NAME_TECHNIQUE], 'F')
            for i in range(len(lstCdTechnique)):
                if len(str(lstTechniqueConcurencyExt[i]).split(",")) > 1:
                    for item in lstTechniqueConcurencyExt[i].split(","):
                        if item != "":
                            left = str(lstCdTechnique[i]).replace(" ", "")
                            right = str(item).replace(" ", "")
                            relationship = Relationship(const.NAME_TECHNIQUE,left, right,const.NAME_TECHNIQUE, const.RELATIONSHIP_TT_EXT)
                            lstRelationship.append(relationship)
                else:
                    if lstTechniqueConcurencyExt[i] is not None:
                        relationship = Relationship(const.NAME_TECHNIQUE,str(lstCdTechnique[i]).strip(), str(lstTechniqueConcurencyExt[i]).strip(),const.NAME_TECHNIQUE, const.RELATIONSHIP_TT_EXT)
                        lstRelationship.append(relationship)
            wb.close()
            return lstRelationship
        except Exception as e :
            common.writeLog("Validate > getRelationshipSheetTechnique() " + str(e))

    # function get data relationship of sheet variable value
    def getRelationshipSheetVariableValue(self):
        try :
            lstRelationship = []
            wb = load_workbook(self.url)
            lstCdVariableType = self.getValuesColumn(wb[const.SHEET_NAME_VARIABLE_VALUE], 'H')

            lstCdVariable = self.getValuesColumn(wb[const.SHEET_NAME_VARIABLE], 'A')
            #create relationship variable value with variable

            lstCdVariableValue = self.getValuesColumn(wb[const.SHEET_NAME_VARIABLE_VALUE], 'B')
            for i in range(len(lstCdVariableType)) :
                if str(lstCdVariableType[i]).strip() in lstCdVariable :
                    relationship = Relationship(const.NAME_VARIABLE_VALUE, str(lstCdVariableValue[i]).strip(),
                                                str(lstCdVariableType[i]).strip(),const.NAME_VARIABLE ,
                                                const.RELATIONSHIP_VARIABLEVALUE_VARIABLE)
                else :

                    relationship = Relationship(const.NAME_VARIABLE_VALUE, str(lstCdVariableValue[i]).strip(),
                                                str(lstCdVariableType[i]).strip(), const.NAME_VARIABLE_VALUE,
                                                const.RELATIONSHIP_VARIABLEVALUE_VARIABLEVALUE)
                lstRelationship.append(relationship)

            wb.close()
            return lstRelationship
        except Exception as e :
            common.writeLog("Validate > getRelationshipSheetVariableValue() " + str(e))

    # function get data relationship of sheet Student
    def getRelationshipSheetStudent(self):
        try:
            lstRelationship = []
            wb = load_workbook(self.url)
            lstGroupCd = self.getValuesColumn(wb[const.SHEET_NAME_STUDENT], 'C')
            lstStudentCd = self.getValuesColumn(wb[const.SHEET_NAME_STUDENT], 'A')
            for i in range(len(lstStudentCd)):
                relationship = Relationship(const.NAME_STUDENT, lstStudentCd[i], lstGroupCd[i], const.NAME_GROUP, const.RELATIONSHIP_STUDENT_GROUP)
                lstRelationship.append(relationship)

            wb.close()
            return lstRelationship
        except Exception as e:
            common.writeLog("Validate > getRelationshipSheetVariableValue() " + str(e))

    # function get data relationship of sheet LocalInsInfo
    def getRelationshipSheetLocalInsInfo(self):
        try:
            lstRelationship = []
            wb = load_workbook(self.url)

            #create relationship student with localinsinfo
            lstLocalInsInfoCd = self.getValuesColumn(wb[const.SHEET_NAME_LOCALINSINFO], 'A')
            lstStudentCd = self.getValuesColumn(wb[const.SHEET_NAME_LOCALINSINFO], 'B')
            for i in range(len(lstLocalInsInfoCd)):
                relationship = Relationship(const.NAME_LOCALINSINFO, lstLocalInsInfoCd[i], lstStudentCd[i], const.NAME_STUDENT, const.RELATIONSHIP_LOCALINSINFO_STUDENT)
                lstRelationship.append(relationship)

            # create relationship technique with localInsInfo
            lstTech = self.getValuesColumn(wb[const.SHEET_NAME_LOCALINSINFO], 'C')
            for i in range(len(lstLocalInsInfoCd)):
                techs = lstTech[i].replace(" ","").split(',')

                if(len(techs) > 1) :
                    for item in techs :
                        relationship = Relationship(const.NAME_LOCALINSINFO,lstLocalInsInfoCd[i], item, const.NAME_TECHNIQUE, const.RELATIONSHIP_LOCALINSINFO_TECHNIQUE)
                        lstRelationship.append(relationship)
                else :
                    relationship = Relationship(const.NAME_LOCALINSINFO, lstLocalInsInfoCd[i], lstTech[i], const.NAME_TECHNIQUE, const.RELATIONSHIP_LOCALINSINFO_TECHNIQUE)
                    lstRelationship.append(relationship)

            # create relationship level with localinsinfo
            sheetLevel = self.getDataSheetLevel()
            lstLevel = self.getValuesColumn(wb[const.SHEET_NAME_LOCALINSINFO], 'D')
            for i in range(len(lstLocalInsInfoCd)):
                levelValue = lstLevel[i]
                levelCd = ""
                for item in sheetLevel:
                    if(item.value == levelValue) :
                        levelCd = item.code
                if( levelCd != "") :
                    relationship = Relationship(const.NAME_LOCALINSINFO, lstLocalInsInfoCd[i], levelCd, const.NAME_LEVEL, const.RELATIONSHIP_LOCALINSINFO_LEVEL)
                    lstRelationship.append(relationship)

            wb.close()
            return lstRelationship
        except Exception as e:
            common.writeLog("Validate > getRelationshipSheetVariableValue() " + str(e))

    # function return data after load excel behind format
    # return data{'Node': {'Variable':[], 'VariableValue':[], 'TaskType': [], 'Technique': []}, 'Relationship' : []}
    def readExcel(self):
        try:
            data = {}
            # create node
            data['Node'] = {const.NAME_VARIABLE : self.getDataSheetVariable(),
                            const.NAME_VARIABLE_VALUE : self.getDataSheetVariableValue(),
                            const.NAME_TASK_TYPE : self.getDataSheetTaskType(),
                            const.NAME_TECHNIQUE : self.getDataSheetTechnique(),
                            const.NAME_LEVEL : self.getDataSheetLevel(),
                            const.NAME_STUDENT : self.getDataSheetStudent(),
                            const.NAME_GROUP : self.getDataSheetGroup(),
                            const.NAME_LOCALINSINFO : self.getDataSheetLocalInsInfo()}

            # create relationship
            data['Relationship'] = self.getRelationshipSheetTaskType() + self.getRelationshipSheetTechnique() \
                                   + self.getRelationshipSheetVariableValue() + self.getRelationshipSheetStudent() \
                                   + self.getRelationshipSheetLocalInsInfo()
            return data
        except Exception as e :
            common.writeLog("Validate > readExcel() " + str(e))




